import os,sys
#os.environ['KERAS_BACKEND'] = 'theano'
import numpy as np
import math
from math import log, pi
import xlrd
import time
import warnings
import pylab
import matplotlib.pyplot as plt
import matplotlib as mpl
import matplotlib.mlab as ml
from scipy import optimize,stats
import pandas as pd
from openpyxl import load_workbook

warnings.simplefilter("ignore",RuntimeWarning)
from random import randint, random

import DataIO
from CoolProp.CoolProp import PropsSI

plt.style.use('Elsevier.mplstyle')
mpl.rcParams['legend.numpoints'] = 1

#--------------------------------------------------------------------------
def Import(start,end,filename):
    "import experimental data"
    
    [data,rownum] = DataIO.ParameterImport(start,end,filename)
    
    i = 0  
    "initialize arrays"
    Tamb = float(data[i][0])
    Tsuc = float(data[i][1])
    Pev = float(data[i][2])
    Tex = float(data[i][3])
    Pcd = float(data[i][4])
    Tinj = float(data[i][5])
    Pinj = float(data[i][6])    
    mgas = float(data[i][7])
    minj = float(data[i][8])
    ratio_minj = float(data[i][9])
    Q = float(data[i][10]) #Q
    W = float(data[i][11])
    eta_s = float(data[i][12])
    i=i+1
    
    while i < (end - start+1):
        Tamb = np.append(Tamb,float(data[i][0]))
        Tsuc = np.append(Tsuc,float(data[i][1]))
        Pev = np.append(Pev,float(data[i][2]))
        Tex = np.append(Tex,float(data[i][3]))
        Pcd = np.append(Pcd,float(data[i][4]))
        Tinj = np.append(Tinj,float(data[i][5]))
        Pinj = np.append(Pinj,float(data[i][6]))        
        mgas = np.append(mgas,float(data[i][7]))
        minj = np.append(minj,float(data[i][8]))
        ratio_minj = np.append(ratio_minj,float(data[i][9]))
        Q = np.append(Q,float(data[i][10]))
        W = np.append(W,float(data[i][11]))
        eta_s = np.append(eta_s,float(data[i][12]))
#        print "i: ",i
        i=i+1
        Data = [Tamb,Tsuc,Pev,Tex,Pcd,Tinj,Pinj,mgas,minj,ratio_minj,Q,W,eta_s]
    
    return Data
    
def Error(W_calc,M_calc,T_calc,W_meas,M_meas,T_meas):
    
    e_W = np.sum(np.square(np.divide(np.subtract(W_calc,W_meas),W_calc)))
    e_M = np.sum(np.square(np.divide(np.subtract(M_calc,M_meas),M_calc)))
    e_T = np.sum(np.square(np.divide(np.subtract(T_calc,T_meas),60)))  #not divided by absolute temp to give eror more weight

    E = math.sqrt(e_W) + math.sqrt(e_M) + math.sqrt(e_T)
    
    return E    


def Normalize(y_data,y_data_min,y_data_max):
    
    y_norm = 0.8*(y_data - y_data_min)/(y_data_max - y_data_min) + 0.1
    
    return y_norm

def DeNormalize(y_norm,y_data_min,y_data_max):
    
    y = (y_norm - 0.1)*(y_data_max - y_data_min)/0.8 + y_data_min
    
    return y


def REmean(y_true,y_pred):
    
    return np.mean(np.fabs(y_true - y_pred)/y_true)    


def Rsquared(y_true,y_pred):
    
    slope, intercept, r_value, p_value, std_err = stats.linregress(y_true,y_pred)    
    
    return r_value**2
    
def Calculate():

    "Import Experimental Data"
    start=1
    end=43
    filename = 'scroll_inj_R407C.csv'
    [T_amb,T_suc,P_ev,T_ex_meas,P_cd,T_inj,P_inj,m_gas_meas,m_inj_meas,ratio_minj,Q_meas,W_meas,eta_s_meas] = Import(start,end,filename)
    
    gas = 'R407C'

    mode = 'training'
    
    from keras.models import Model,Sequential,load_model
    from keras.layers import Input,Flatten,Dense,LSTM,merge,Dropout,concatenate
    #from keras.engine import merge # from Keras version 1.2.2
    #from keras.layers.merge import concatenate
    from keras.utils import plot_model
    from keras.callbacks import TensorBoard
    from keras import regularizers
    

    #Define inputs
    P_ev = P_ev.reshape(-1,1,1)
    T_suc = T_suc.reshape(-1,1,1)
    P_cd = P_cd.reshape(-1,1,1)
    T_inj = T_inj.reshape(-1,1,1)
    P_inj = P_inj.reshape(-1,1,1)
    T_amb = T_amb.reshape(-1,1,1)

    P_ev_norm = Normalize(P_ev,100, 900)
    T_suc_norm = Normalize(T_suc,263.15,300)
    P_cd_norm = Normalize(P_cd,1000, 3500)
    T_inj_norm = Normalize(T_inj,273.15,330.15)
    P_inj_norm = Normalize(P_inj,300, 1700)
    T_amb_norm = Normalize(T_amb, 290.15, 330.15)
    m_gas_meas_norm = Normalize(m_gas_meas,0.01,0.2)
    m_inj_meas_norm = Normalize(m_inj_meas,0.001,0.05)
    W_meas_norm = Normalize(W_meas, 1000,8000)
    T_ex_meas_norm = Normalize(T_ex_meas,310,390)
    eta_s_meas_norm = Normalize(eta_s_meas,0.3,0.8)
    Q_meas_norm = Normalize(Q_meas,-30,500)
    
    if mode == 'training':
        visible1 = Input(shape=(1,1), name='P_ev')
        visible2 = Input(shape=(1,1), name='T_suc')
        visible3 = Input(shape=(1,1), name='P_cd')
        visible4 = Input(shape=(1,1), name='T_inj')
        visible5 = Input(shape=(1,1), name='P_inj')
        visible6 = Input(shape=(1,1), name='T_amb')
    
        shared_lstm = LSTM(100)
    
        encoded_a = shared_lstm(visible1)
        encoded_b = shared_lstm(visible2)
        encoded_c = shared_lstm(visible3)
        encoded_d = shared_lstm(visible4)
        encoded_e = shared_lstm(visible5)
        encoded_f = shared_lstm(visible6)
    
        #Merge inputs
        #merged = merge([encoded_a,encoded_b,encoded_c,encoded_d,encoded_e,encoded_f],mode='concat',concat_axis=-1) #deprecated
        merged = concatenate([encoded_a,encoded_b,encoded_c,encoded_d,encoded_e,encoded_f],axis=-1)
        
        #interpretation model
        hidden1 = Dense(256, activation='tanh')(merged)
        hidden2 = Dense(128, activation = 'tanh')(hidden1)
        #hidden3 = Dropout(0.2, noise_shape=None, seed=None)(hidden2)
        hidden3 = Dense(64, activation = 'tanh')(hidden2)
        output1 = Dense(1, activation = 'linear',name='m_gas')(hidden3)
        output2 = Dense(1, activation = 'linear',name='m_inj')(hidden3)
        output3 = Dense(1, activation = 'linear',name='W_dot')(hidden3)
        output4 = Dense(1, activation = 'linear',name='T_ex')(hidden3)
        output5 = Dense(1, activation = 'linear',name='eta_s')(hidden3)
        output6 = Dense(1, activation = 'tanh',name='f_q')(hidden3)
       
        model = Model(input=[visible1,visible2,visible3,visible4,visible5,visible6],
                        output = [output1,output2,output3,output4,output5,output6])
        
        plot_model(model, to_file='model.png',show_shapes=True,show_layer_names=True)
        
        model.compile(optimizer='adamax',loss=['mse','mse','mse','mse','mse','mse'])
        
        history = model.fit([P_ev_norm,T_suc_norm,P_cd_norm,T_inj_norm,P_inj_norm,T_amb_norm],
                            [m_gas_meas_norm,m_inj_meas_norm,W_meas_norm,T_ex_meas_norm,eta_s_meas_norm,Q_meas_norm],
                            epochs=8000 ,
                            batch_size=30,
                            validation_split=0.2,
                            )
    
    #   #History plot
        fig=pylab.figure(figsize=(3.5,2.5))
        plt.semilogy(history.history['loss'])
        plt.semilogy(history.history['val_loss'])
        plt.ylabel('loss [-]')
        plt.xlabel('epoch [-]')
        plt.legend(['train', 'test'], loc='upper right')
        #plt.ylim(0,0.1)
        plt.tight_layout(pad=0.2)  
        plt.tick_params(direction='in')      
        fig.savefig('ANN_history_scroll_inj_R407C.pdf')    

        # Save the model
        model.save('ANN_model_scroll_inj_combined.h5')
    
    elif mode == 'run':
    
        # Load the model
        model = load_model('ANN_model_scroll_inj.h5')
    
    # Run the model
    [Mref,Minj,W,T,eta_s,Q] = model.predict([P_ev_norm,T_suc_norm,P_cd_norm,T_inj_norm,P_inj_norm,T_amb_norm])
    W = DeNormalize(W.reshape(-1),1000,8000)
    Mref = DeNormalize(Mref.reshape(-1),0.01,0.2)
    Minj = DeNormalize(Minj.reshape(-1),0.001,0.05)
    T = DeNormalize(T.reshape(-1),310,390)
    eta_s = DeNormalize(eta_s.reshape(-1),0.3,0.8)
    Q = DeNormalize(Q.reshape(-1),-30,500)


    # Save the architecture of a model, and not its weights or its training configuration
    # save as JSON
    # json_string = model.to_json()
    
    # save as YAML
    # yaml_string = model.to_yaml()


    for i in range(0,(end-start+1)):


        data_calc = {'Tdis':[T[i]],'mdot':[Mref[i]],'mdot_inj':[Minj[i]], 'Wdot':[W[i]],'etaoa':[eta_s[i]],'fq':[Q[i]/W[i]]} 
            
        
        # Write to Excel
        filename = os.path.dirname(__file__)+'/VIScroll_output.xlsx'
        xl = pd.read_excel(filename, sheet_name='ANN_Validation')

        df = pd.DataFrame(data=data_calc)

        df.reindex(columns=xl.columns)
        df_final=xl.append(df,ignore_index=True)
        df_final.tail()
        
        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl',index=False)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df_final.to_excel(writer,index=False,sheet_name='ANN_Validation')
        
        # 
        writer.save()

    
    #Separate testing from calibrating
    n_len = len(W)
    n_split = int(np.floor(0.3*n_len))
    n_training = int(n_len-n_split-1)

    # Validation power output
    fig=pylab.figure(figsize=(3.5,2.5))

    plt.plot(W[:n_training]/1000,W_meas[:n_training]/1000,'ro',ms=3,mec='black',mew=0.5, label='training points')
    plt.plot(W[-n_split:]/1000,W_meas[-n_split:]/1000,'b*',ms=4,mec='black',mew=0.5,label='testing points')

    plt.ylabel('Input Power (calc) [kW]',fontsize = 10)
    plt.xlabel('Input Power (meas) [kW]',fontsize = 10)
    Wmin = 0
    Wmax = 5000
    x=[Wmin/1000,Wmax/1000]
    y=[Wmin/1000,Wmax/1000]
    y105=[1.02*Wmin/1000,1.02*Wmax/1000]
    y95=[0.98*Wmin/1000,0.98*Wmax/1000]
    
    plt.plot(x,y,'k-')
    plt.fill_between(x,y105,y95,color='black',alpha=0.2)
    plt.xlim(3,5)
    plt.ylim(3,5)
    plt.legend(loc=2, fontsize=9)
    plt.tick_params(direction='in')
    plt.tight_layout(pad=0.2)        
    #plt.show()
    fig.savefig('ANN_W_scroll_inj.pdf')

    # Validation heat loss
    fig=pylab.figure(figsize=(3.5,2.5))

    plt.plot(Q[:n_training]/W[:n_training],Q_meas[:n_training]/W_meas[:n_training],'ro',ms=3,mec='black',mew=0.5,label='training points')
    plt.plot(Q[-n_split:]/W[-n_split:],Q_meas[-n_split:]/W_meas[-n_split:],'b*',ms=4,mec='black',mew=0.5,label='testing points')
    
    plt.ylabel('Heat Loss Fraction (calc) [-]',fontsize = 10)
    plt.xlabel('Heat Loss Fraction (meas) [-]',fontsize = 10)
    fqmin = 0
    fqmax = 0.1
    x=[fqmin,fqmax]
    y=[fqmin,fqmax]
    y105=[1.05*fqmin,1.05*fqmax]
    y95=[0.95*fqmin,0.95*fqmax]
    
    plt.plot(x,y,'k-')
    plt.fill_between(x,y105,y95,color='black',alpha=0.2)
    plt.xlim(0,0.1)
    plt.ylim(0,0.1)
    plt.legend(loc=2,fontsize=9)
    plt.tick_params(direction='in')
    plt.tight_layout(pad=0.2)        
    #plt.show()
    fig.savefig('ANN_fq_scroll_inj.pdf')


    # Validation mass flow rate
    fig=pylab.figure(figsize=(3.5,2.5))
    plt.plot(Mref[:n_training],m_gas_meas[:n_training],'ro',ms = 3,mec='black',mew=0.5,label='training points')
    plt.plot(Mref[-n_split:],m_gas_meas[-n_split:],'b*',ms = 4,mec='black',mew=0.5,label='testing points')
    
    plt.xlabel('Mass flow rate (calc) [kg/s]',fontsize = 10)
    plt.ylabel('Mass flow rate (meas) [kg/s]',fontsize = 10)
    
    Mmin = 0.04
    Mmax = 0.1
    x=[Mmin,Mmax]
    y=[Mmin,Mmax]
    y105=[1.05*Mmin,1.05*Mmax]
    y95=[0.95*Mmin,0.95*Mmax]
    
    plt.plot(x,y,'k-')
    plt.fill_between(x,y105,y95,color='black',alpha=0.2)
    plt.plot(loc=2,fontsize=9)
    plt.xlim(Mmin,Mmax)
    plt.ylim(Mmin,Mmax)
    plt.legend(loc=2,fontsize=9)
    plt.tick_params(direction='in')
    plt.tight_layout(pad=0.2)        
    #plt.show()
    fig.savefig('ANN_M_scroll_inj.pdf')

    # Validation injection mass flow rate
    fig=pylab.figure(figsize=(3.5,2.5))
    plt.plot(Minj[:n_training]*1000,m_inj_meas[:n_training]*1000,'ro',ms = 3,mec='black',mew=0.5,label='training points')
    plt.plot(Minj[-n_split:]*1000,m_inj_meas[-n_split:]*1000,'b*',ms = 4,mec='black',mew=0.5,label='testing points')
    
    plt.xlabel('Inj Mass flow rate (calc) [g/s]',fontsize = 10)
    plt.ylabel('Inj Mass flow rate (meas) [g/s]',fontsize = 10)
    
    Mmin = 4
    Mmax = 30
    x=[Mmin,Mmax]
    y=[Mmin,Mmax]
    y105=[1.05*Mmin,1.05*Mmax]
    y95=[0.95*Mmin,0.95*Mmax]
    
    plt.plot(x,y,'k-')
    plt.fill_between(x,y105,y95,color='black',alpha=0.2)
    plt.legend(loc=2,fontsize=9)
    plt.xlim(Mmin,Mmax)
    plt.ylim(Mmin,Mmax)
    plt.tick_params(direction='in')
    plt.tight_layout(pad=0.2)        
    #plt.show()
    fig.savefig('ANN_Minj_scroll_inj.pdf')


    # Validation discharge temperature
    fig=pylab.figure(figsize=(3.5,2.5))

    plt.plot(T[:n_training],T_ex_meas[:n_training],'ro',ms = 3,mec='black',mew=0.5,label='training points')
    plt.plot(T[-n_split:],T_ex_meas[-n_split:],'b*',ms = 4,mec='black',mew=0.5,label='testing points')

    plt.xlabel('Discharge Temperature (calc) [K]',fontsize = 10)
    plt.ylabel('Discharge Temperature\n(meas) [K]',fontsize = 10)
    
    Tmin = 340
    Tmax = 380
    x=[Tmin,Tmax]
    y=[Tmin,Tmax]
    y105=[2+Tmin,2+Tmax]
    y95=[Tmin-2,Tmax-2]
    
    plt.plot(x,y,'k-')
    plt.fill_between(x,y105,y95,color='black',alpha=0.2)    
    plt.xlim(Tmin,Tmax)
    plt.ylim(Tmin,Tmax)
    plt.legend(loc=2,fontsize=9)
    plt.tight_layout(pad=0.2)        
    plt.tick_params(direction='in')
    #plt.show()
    fig.savefig('ANN_T_scroll_inj.pdf')


    # Validation isentropic efficiency
    
    fig=pylab.figure(figsize=(3.5,2.5))
    plt.plot(eta_s_meas[:n_training],eta_s[:n_training],'ro',ms = 3,mec='black',mew=0.5,label='training points')
    plt.plot(eta_s_meas[-n_split:],eta_s[-n_split:],'b*',ms = 4,mec='black',mew=0.5,label='testing points')

    plt.ylabel('Isentropic Efficiency (calc) [-]',fontsize = 10)
    plt.xlabel('Isentropic Efficiency (meas) [-]',fontsize = 10)
    
    eta_min = 0
    eta_max = 0.7
    x=[eta_min,eta_max]
    y=[eta_min,eta_max]
    y105=[1.02*eta_min,1.02*eta_max]
    y95=[0.98*eta_min,0.98*eta_max]
    
    plt.plot(x,y,'k-')
    plt.fill_between(x,y105,y95,color='black',alpha=0.2)
    plt.legend(loc=2,fontsize=9)
    plt.xlim(0.6,0.7)
    plt.ylim(0.6,0.7)
    plt.tick_params(direction='in')
    plt.tight_layout(pad=0.2)
    fig.savefig('ANN_etas_scroll_inj.pdf')


    
    #err = Error(W,M,T,W_dot_meas,m_dot_meas,T_ex_meas)
    #print err
    print 'Wdot:',REmean(W_meas,W),Rsquared(W_meas,W)*100
    print 'mdot:',REmean(m_gas_meas,Mref),Rsquared(m_gas_meas,Mref)*100
    print 'minj:',REmean(m_inj_meas,Minj),Rsquared(m_inj_meas,Minj)*100
    print 'Tex:',REmean(T_ex_meas,T),Rsquared(T_ex_meas,T)*100
    print 'eta_s:',REmean(eta_s_meas,eta_s),Rsquared(eta_s_meas,eta_s)*100
    print 'f_q:',REmean(Q_meas,Q),Rsquared(Q_meas,Q)*100
    print 'm_inj/m_suc:',REmean(m_inj_meas/m_gas_meas,Minj/Mref),Rsquared(m_inj_meas/m_gas_meas,Minj/Mref)*100
    



def Validation_Ammar_2phase():

    "Import Experimental Data"
    start=2
    end=12
    filename = 'scroll_inj_R407C_2ph.csv'
    [T_amb,T_suc,P_ev,T_ex_meas,P_cd,T_inj,P_inj,m_gas_meas,m_inj_meas,ratio_minj,f_q_meas,W_meas,eta_s_meas] = Import(start,end,filename)
    
    gas = 'R407C'


    from keras.models import Model,Sequential,load_model
    from keras.layers import Input,Flatten,Dense,LSTM,merge
    #from keras.engine import merge # from Keras version 1.2.2
    from keras.layers.merge import concatenate
    from keras.utils import plot_model
    from keras.callbacks import TensorBoard
    

    #Define inputs
    P_ev = P_ev.reshape(-1,1,1)
    T_suc = T_suc.reshape(-1,1,1)
    P_cd = P_cd.reshape(-1,1,1)
    T_inj = T_inj.reshape(-1,1,1)
    P_inj = P_inj.reshape(-1,1,1)
    T_amb = T_amb.reshape(-1,1,1)

    P_ev_norm = Normalize(P_ev,100, 900)
    T_suc_norm = Normalize(T_suc,263.15,300)
    P_cd_norm = Normalize(P_cd,1000, 3500)
    T_inj_norm = Normalize(T_inj,273.15,330.15)
    P_inj_norm = Normalize(P_inj,300, 1700)
    T_amb_norm = Normalize(T_amb, 290.15, 330.15)
    m_gas_meas_norm = Normalize(m_gas_meas,0.01,0.2)
    m_inj_meas_norm = Normalize(m_inj_meas,0.001,0.05)
    W_meas_norm = Normalize(W_meas, 1000,8000)
    T_ex_meas_norm = Normalize(T_ex_meas,310,390)
    eta_s_meas_norm = Normalize(eta_s_meas,0.3,0.8)
    f_q_meas_norm = Normalize(f_q_meas,-30,500)


    # Load the model
    model = load_model('ANN_model_scroll_inj_combined.h5')
    
    # Run the model
    [Mref,Minj,W,T,eta_s,f_q] = model.predict([P_ev_norm,T_suc_norm,P_cd_norm,T_inj_norm,P_inj_norm,T_amb_norm])
    W = DeNormalize(W.reshape(-1),1000,8000)
    Mref = DeNormalize(Mref.reshape(-1),0.01,0.2)
    Minj = DeNormalize(Minj.reshape(-1),0.001,0.05)
    T = DeNormalize(T.reshape(-1),310,390)
    eta_s = DeNormalize(eta_s.reshape(-1),0.3,0.8)
    f_q = DeNormalize(f_q.reshape(-1),-30,500)


    for i in range(0,(end-start+1)):


        data_calc = {'Tdis':[T[i]],'mdot':[Mref[i]],'mdot_inj':[Minj[i]], 'Wdot':[W[i]],'etaoa':[eta_s[i]],'fq':[f_q[i]/W[i]]} 
            
        
        # Write to Excel
        filename = os.path.dirname(__file__)+'/VIScroll_output.xlsx'
        xl = pd.read_excel(filename, sheet_name='ANN_Validation_2phase')
    
        df = pd.DataFrame(data=data_calc)
    
        df.reindex(columns=xl.columns)
        df_final=xl.append(df,ignore_index=True)
        df_final.tail()
        
        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl',index=False)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df_final.to_excel(writer,index=False,sheet_name='ANN_Validation_2phase')
        
        # 
        writer.save()


    # Validation power output
    fig=pylab.figure(figsize=(3.5,2.5))

    plt.plot(W/1000,W_meas/1000,'ro',ms=3,mec='black',mew=0.5, label='training points')


    plt.ylabel('Input Power (calc) [kW]',fontsize = 10)
    plt.xlabel('Input Power (meas) [kW]',fontsize = 10)
    Wmin = 0
    Wmax = 8000
    x=[Wmin/1000,Wmax/1000]
    y=[Wmin/1000,Wmax/1000]
    y105=[1.02*Wmin/1000,1.02*Wmax/1000]
    y95=[0.98*Wmin/1000,0.98*Wmax/1000]
    
    plt.plot(x,y,'k-')
    plt.fill_between(x,y105,y95,color='black',alpha=0.2)
    plt.xlim(3,8)
    plt.ylim(3,8)

    plt.tick_params(direction='in')
    plt.tight_layout(pad=0.2)        
    #plt.show()
    fig.savefig('ANN_W_scroll_inj_2ph.pdf')

    # Validation heat loss
    fig=pylab.figure(figsize=(3.5,2.5))

    plt.plot(f_q/W,f_q_meas/W_meas,'ro',ms=3,mec='black',mew=0.5,label='training points')

    
    plt.ylabel('Heat Loss Fraction (calc) [-]',fontsize = 10)
    plt.xlabel('Heat Loss Fraction (meas) [-]',fontsize = 10)
    fqmin = 0
    fqmax = 0.1
    x=[fqmin,fqmax]
    y=[fqmin,fqmax]
    y105=[1.05*fqmin,1.05*fqmax]
    y95=[0.95*fqmin,0.95*fqmax]
    
    plt.plot(x,y,'k-')
    plt.fill_between(x,y105,y95,color='black',alpha=0.2)
    plt.xlim(0,0.1)
    plt.ylim(0,0.1)

    plt.tick_params(direction='in')
    plt.tight_layout(pad=0.2)        
    #plt.show()
    fig.savefig('ANN_fq_scroll_inj_2ph.pdf')


    # Validation mass flow rate
    fig=pylab.figure(figsize=(3.5,2.5))
    plt.plot(Mref,m_gas_meas,'ro',ms = 3,mec='black',mew=0.5,label='training points')

    
    plt.xlabel('Mass flow rate (calc) [kg/s]',fontsize = 10)
    plt.ylabel('Mass flow rate (meas) [kg/s]',fontsize = 10)
    
    Mmin = 0.04
    Mmax = 0.1
    x=[Mmin,Mmax]
    y=[Mmin,Mmax]
    y105=[1.05*Mmin,1.05*Mmax]
    y95=[0.95*Mmin,0.95*Mmax]
    
    plt.plot(x,y,'k-')
    plt.fill_between(x,y105,y95,color='black',alpha=0.2)
    plt.plot(loc=2,fontsize=9)
    plt.xlim(Mmin,Mmax)
    plt.ylim(Mmin,Mmax)

    plt.tick_params(direction='in')
    plt.tight_layout(pad=0.2)        
    #plt.show()
    fig.savefig('ANN_M_scroll_inj_2ph.pdf')

    # Validation injection mass flow rate
    fig=pylab.figure(figsize=(3.5,2.5))
    plt.plot(Minj*1000,m_inj_meas*1000,'ro',ms = 3,mec='black',mew=0.5,label='training points')

    
    plt.xlabel('Inj Mass flow rate (calc) [g/s]',fontsize = 10)
    plt.ylabel('Inj Mass flow rate (meas) [g/s]',fontsize = 10)
    
    Mmin = 4
    Mmax = 30
    x=[Mmin,Mmax]
    y=[Mmin,Mmax]
    y105=[1.05*Mmin,1.05*Mmax]
    y95=[0.95*Mmin,0.95*Mmax]
    
    plt.plot(x,y,'k-')
    plt.fill_between(x,y105,y95,color='black',alpha=0.2)

    plt.xlim(Mmin,Mmax)
    plt.ylim(Mmin,Mmax)
    plt.tick_params(direction='in')
    plt.tight_layout(pad=0.2)        
    #plt.show()
    fig.savefig('ANN_Minj_scroll_inj_2ph.pdf')


    # Validation discharge temperature
    fig=pylab.figure(figsize=(3.5,2.5))

    plt.plot(T,T_ex_meas,'ro',ms = 3,mec='black',mew=0.5,label='training points')


    plt.xlabel('Discharge Temperature (calc) [K]',fontsize = 10)
    plt.ylabel('Discharge Temperature\n(meas) [K]',fontsize = 10)
    
    Tmin = 340
    Tmax = 380
    x=[Tmin,Tmax]
    y=[Tmin,Tmax]
    y105=[2+Tmin,2+Tmax]
    y95=[Tmin-2,Tmax-2]
    
    plt.plot(x,y,'k-')
    plt.fill_between(x,y105,y95,color='black',alpha=0.2)    
    plt.xlim(Tmin,Tmax)
    plt.ylim(Tmin,Tmax)

    plt.tight_layout(pad=0.2)        
    plt.tick_params(direction='in')
    #plt.show()
    fig.savefig('ANN_T_scroll_inj_2ph.pdf')


    # Validation isentropic efficiency
    
    fig=pylab.figure(figsize=(3.5,2.5))
    plt.plot(eta_s_meas,eta_s,'ro',ms = 3,mec='black',mew=0.5,label='training points')

    plt.ylabel('Isentropic Efficiency (calc) [-]',fontsize = 10)
    plt.xlabel('Isentropic Efficiency (meas) [-]',fontsize = 10)
    
    eta_min = 0
    eta_max = 0.7
    x=[eta_min,eta_max]
    y=[eta_min,eta_max]
    y105=[1.02*eta_min,1.02*eta_max]
    y95=[0.98*eta_min,0.98*eta_max]
    
    plt.plot(x,y,'k-')
    plt.fill_between(x,y105,y95,color='black',alpha=0.2)

    plt.xlim(0.5,0.7)
    plt.ylim(0.5,0.7)
    plt.tick_params(direction='in')
    plt.tight_layout(pad=0.2)
    fig.savefig('ANN_etas_scroll_inj_2ph.pdf')


    
    #err = Error(W,M,T,W_dot_meas,m_dot_meas,T_ex_meas)
    #print err
    print 'Wdot:',REmean(W_meas,W),Rsquared(W_meas,W)*100
    print 'mdot:',REmean(m_gas_meas,Mref),Rsquared(m_gas_meas,Mref)*100
    print 'minj:',REmean(m_inj_meas,Minj),Rsquared(m_inj_meas,Minj)*100
    print 'Tex:',REmean(T_ex_meas,T),Rsquared(T_ex_meas,T)*100
    print 'eta_s:',REmean(eta_s_meas,eta_s),Rsquared(eta_s_meas,eta_s)*100
    print 'f_q:',REmean(f_q_meas,f_q),Rsquared(f_q_meas,f_q)*100
    print 'm_inj/m_suc:',REmean(m_inj_meas/m_gas_meas,Minj/Mref),Rsquared(m_inj_meas/m_gas_meas,Minj/Mref)*100
    
if __name__ == '__main__':
    
    Calculate()
    #Validation_Ammar_2phase()
    
    """
    Wdot: 0.008220341979717799 99.56922288736178
    mdot: 0.011331759697160361 99.64595206448986
    minj: 0.07011354008261375 96.63843422122608
    Tex: 0.002574871815676044 98.65646548151139
    eta_s: 0.009537828936222303 79.76583704167605
    f_q: 0.07263509578119003 92.44388296110533
    m_inj/m_suc: 0.07515435340478782 96.01705488251565
    """
    
    
    
    
    