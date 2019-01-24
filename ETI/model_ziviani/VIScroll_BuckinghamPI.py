from __future__ import division, print_function, absolute_import
import os,sys
from CoolProp.CoolProp import PropsSI
import CoolProp as CP
import pandas as pd
from openpyxl import load_workbook
import pylab
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.mlab as ml
from math import cos,sqrt
plt.style.use('Elsevier')

def savefigs(name):
    #plt.savefig(name+'.eps')
    plt.savefig(name+'.pdf')
    plt.savefig(name+'.png',dpi=600)


def MAPE(y_true,y_pred):
    """
    Function returning the Mean Absolute Percentage Error
    """

    return np.mean(np.abs((y_true - y_pred) / y_true))*100 
    
def REmean(y_true,y_pred):
    
    return np.mean(np.fabs(y_true - y_pred)/y_true)

def RE(y_true,y_pred):
    """
    Function returning the Maximum Relative Percentage Error
    """

    return np.max(np.abs((y_true - y_pred) / y_true))*100 

class VICompressorLumpkinBPIClass():

    """


    

    Required Parameters:

        

    ===========   ==========  ========================================================================

    Variable      Units       Description

    ===========   ==========  ========================================================================

    M             Ibm/hr      A numpy-like list of compressor map coefficients for mass flow

    P             Watts       A numpy-like list of compressor map coefficients for electrical power

    Ref           N/A         A string representing the refrigerant

    Tin_r         K           Refrigerant inlet temperature

    pin_r         Pa          Refrigerant suction pressure (absolute)

    pout_r        Pa          Refrigerant discharge pressure (absolute)

    fp            --          Fraction of electrical power lost as heat to ambient

    Vdot_ratio    --          Displacement Scale factor

    ===========   ==========  ========================================================================

    

    All variables are of double-type unless otherwise specified

        

    """

    def __init__(self,**kwargs):

        #Load up the parameters passed in
        # using the dictionary
        self.__dict__.update(kwargs)

    def Update(self,**kwargs):

        #Update the parameters passed in
        # using the dictionary
        self.__dict__.update(kwargs)

        

    def OutputList(self):

        """
            Return a list of parameters for this component for further output

            It is a list of tuples, and each tuple is formed of items with indices:

                [0] Description of value
                [1] Units of value
                [2] The value itself

        """

        return [ 
            ('Heat Loss Fraction','-',self.fp),
            ('Displacement scale factor','-',self.Vdot_ratio),
            ('Power','W',self.W),
            ('Suction mass flow rate','kg/s',self.mdot_r),
            ('Injection mass flow rate','kg/s',self.mdot_inj),
            ('Total mass flow rate','kg/s',self.mdot_tot),
            ('Inlet Temperature','K',self.Tin_r),
            ('Injection Temperature','K',self.Tinj_r),
            ('Outlet Temperature','K',self.Tout_r),
            ('Inlet Enthalpy','J/kg',self.hin_r),
            ('Injection Enthalpy','J/kg',self.hinj_r),
            ('Outlet Enthalpy','J/kg',self.hout_r),
            ('Inlet Pressure','Pa',self.pin_r),
            ('Injection Pressure','Pa',self.pinj_r),
            ('Outlet Pressure','Pa',self.pout_r),
            ('Inlet Entropy','J/kg-K',self.sin_r),
            ('Injection Entropy','J/kg-K',self.sinj_r),
            ('Outlet Entropy','J/kg-K',self.sout_r),
            ('Overall isentropic efficiency','-',self.eta_oi),
            ('Pumped flow rate','m**3/s',self.Vdot_pumped),
            ('Ambient heat loss','W',self.Q_amb)
         ]

        

    def Calculate(self):

        #AbstractState
        AS = self.AS

        #Local copies of coefficients
        P=self.P
        R=self.R
        T=self.T
        V=self.V
        F=self.F
        IS=self.IS

        #Critical temperature and pressure
        Tcrit = AS.T_critical() #[K]
        pcrit = AS.p_critical() #[Pa]

        #Calculate suction and injection dew temperatures
        AS.update(CP.PQ_INPUTS, self.pin_r, 0.0)
        h_bubble=AS.hmass() #[J/kg]

        AS.update(CP.PQ_INPUTS, self.pinj_r, 0.0)
        h_bubble_inj=AS.hmass() #[J/kg]

        AS.update(CP.PT_INPUTS, self.pinj_r, self.Tinj_r)
        self.hinj_r=AS.hmass() #[J/kg]
        self.sinj_r=AS.smass() #[J/kg-K]

        AS.update(CP.PT_INPUTS, self.pin_r, self.Tin_r)
        self.hin_r=AS.hmass() #[J/kg]
        self.sin_r=AS.smass() #[J/kg-K]
        self.vin_r = 1 / AS.rhomass() #[m**3/kg]

        #AS.update(CP.QT_INPUTS, 0.0, self.Tinj_r)
        AS.update(CP.PQ_INPUTS, self.pinj_r, 0.0)
        h_f_inj=AS.hmass()#[J/kg]

        #AS.update(CP.QT_INPUTS, 1.0, self.Tinj_r)
        AS.update(CP.PQ_INPUTS, self.pinj_r, 1.0)
        h_g_inj=AS.hmass()#[J/kg]

        #AS.update(CP.QT_INPUTS, 0.0, self.Tin_r)
        AS.update(CP.PQ_INPUTS, self.pin_r, 0.0)
        h_f=AS.hmass()#[J/kg]

        #AS.update(CP.QT_INPUTS, 1.0, self.Tin_r)
        AS.update(CP.PQ_INPUTS, self.pin_r, 1.0)
        h_g=AS.hmass()#[J/kg]

        #Injection:
        Dh_shinj_r=self.hinj_r-h_bubble_inj
        Dh_fg_inj=h_g_inj-h_f_inj

        #Suction
        Dh_sh=self.hin_r-h_bubble
        Dh_fg=h_g-h_f
        
        #Normalized quantities
        f_norm = self.f_power/self.f_nominal
        p_inj_norm = self.pinj_r/self.pin_r
        deltah_inj_norm = Dh_shinj_r/Dh_fg_inj
        p_dis_norm = self.pout_r/pcrit
        p_suc_norm = self.pin_r/pcrit
        p_ratio = self.pout_r/self.pin_r
        deltah_suc_norm = Dh_sh/Dh_fg
        T_suc_norm = self.Tin_r/self.Tamb
        T_suc2_norm = self.Tin_r/Tcrit

        #Ratio of injection to suction mass flow rate [-]
        ratio_mass = R[0]*(f_norm**R[1])*(deltah_inj_norm**R[2])*(p_inj_norm**R[3])*(T_suc_norm**R[4])*(p_ratio**R[5])*(deltah_suc_norm**R[6])*(p_dis_norm**R[7])*(p_suc_norm**R[8])*(T_suc2_norm**R[9])

        #Volumetric efficiency
        eta_v = V[0]*f_norm**V[1]*ratio_mass**V[2]*p_inj_norm**V[3]*T_suc_norm**V[4]*p_ratio**V[5]*deltah_suc_norm**V[6]*p_dis_norm**V[7]*p_suc_norm**V[8]*T_suc2_norm**V[9]
        
        #Suction mass flow rate [kg/s]
        mdot = eta_v*self.Vdisp*(self.f_power*60)/self.vin_r

        #Injection mass flow rate [kg/s]
        mdot_inj = mdot*ratio_mass

        #Discharge tempearture [K]
        T_dis = Tcrit*(T[0]*f_norm**T[1]*ratio_mass**T[2]*p_inj_norm**T[3]*T_suc_norm**T[4]*p_ratio**T[5]*deltah_suc_norm**T[6]*p_dis_norm**T[7]*p_suc_norm**T[8]*T_suc2_norm**T[9])

        #Power [Watts]
        power = self.Wdot_max*(P[0]*f_norm**P[1]*ratio_mass**P[2]*p_inj_norm**P[3]*T_suc_norm**P[4]*p_ratio**P[5]*deltah_suc_norm**P[6]*p_dis_norm**P[7]*p_suc_norm**P[8]*T_suc2_norm**P[9])

        #Heat loss fraction [-]
        fq = F[0]*f_norm**F[1]*ratio_mass**F[2]*p_inj_norm**F[3]*T_suc_norm**F[4]*p_ratio**F[5]*deltah_suc_norm**F[6]*p_dis_norm**F[7]*p_suc_norm**F[8]*T_suc2_norm**F[9]

        #Isentropic efficiency
        eta_is = IS[0]*f_norm**IS[1]*ratio_mass**IS[2]*p_inj_norm**IS[3]*T_suc_norm**IS[4]*p_ratio**IS[5]*deltah_suc_norm**IS[6]*p_dis_norm**IS[7]*p_suc_norm**IS[8]*T_suc2_norm**IS[9]
        
        # Discharge state
        self.Tout_r = T_dis #[K]
        AS.update(CP.PT_INPUTS, self.pout_r, self.Tout_r)
        self.hout_r = AS.hmass() #[J/kg]
        self.sout_r = AS.smass() #[J/kg-K]        

        
        #define properites for isentropic efficency
        AS.update(CP.PSmass_INPUTS, self.pout_r, self.sin_r)
        h_2s=AS.hmass() #[J/kg]

        AS.update(CP.PSmass_INPUTS, self.pout_r, self.sinj_r)
        h_4s=AS.hmass() #[J/kg]

        #isentropic effeicincy defined by Groll
        self.eta_oi=eta_is #(mdot*(h_2s-self.hin_r) + mdot_inj*(h_4s-self.hinj_r))/power

        # Output
        self.mdot_r = mdot
        self.mdot_inj = mdot_inj
        self.mdot_tot = mdot + mdot_inj
        self.W=power
        self.eta_v = eta_v
        self.CycleEnergyIn=power*(1-self.fp)
        self.Vdot_pumped= mdot*self.vin_r
        self.fq = fq
        self.Q_amb=-fq*power

def Validate_VaporInjection():
 
    # Read excel file scroll compressorr
    filename = os.path.dirname(__file__)+'\VIScroll_output.xlsx'
    df_meas = pd.read_excel(open(filename,'rb'), sheet_name='Experimental_Data_SH',skip_footer=0)

    Tamb = np.asarray(df_meas['Tamb'].values.T[:].tolist()) #[K]
    Tsuc = np.asarray(df_meas['Tsuc'].values.T[:].tolist()) #[K]
    psuc = np.asarray(df_meas['Pev'].values.T[:].tolist())*1000 #[Pa]
    Tdis = np.asarray(df_meas['Tdis'].values.T[:].tolist()) #[K]
    pdis = np.asarray(df_meas['Pcond'].values.T[:].tolist())*1000 #[Pa]
    Tinj = np.asarray(df_meas['Tinj'].values.T[:].tolist()) #[K]
    pinj = np.asarray(df_meas['Pinj'].values.T[:].tolist())*1000 #[Pa]

    for i in range(len(Tamb)):
    
        #Abstract State        
        Ref = 'R407C'
        Backend = 'HEOS' #choose between: 'HEOS','TTSE&HEOS','BICUBIC&HEOS','REFPROP','SRK','PR'
        AS = CP.AbstractState(Backend, Ref)
        

        kwds={
            'P':[2.671,1,0.06768,0.08695,0.4397,0.4899,-0.6598,0.1848,0.699,0.3221],
            'R':[0.08758,1,-1.97,0.5076,-2.745,1.334,-1.47,0.6262,0.3004,-0.2136],
            'T':[1.206,1,-0.01493,-0.005106,-0.5655,0.4426,-0.7388,-0.2694,0.2932,1.474],
            'V':[1.169,1,0.00497,-0.02236,-0.6345,0.3229,-0.1655,-0.3987,0.2806,1.856],
            'IS':[0.4,1,0.02304,-0.3125,3.792,0.5774,0.3489,-0.4794,-0.07093,1.974],
            'F':[10,1,0.3288,-1.248,6.196,0.9401,2.269,0.3922,0.4534,-2.865],
            'AS':AS,
            'Tin_r':Tsuc[i],
            'pin_r':psuc[i],
            'pout_r':pdis[i],
            'pinj_r':pinj[i],
            'Tinj_r':Tinj[i],
            'Tamb':Tamb[i],
            'Vdisp': 67.186962e-6, #[m3/rev]
            'f_power': 60, #[Hz]
            'f_nominal': 60, #[Hz]
            'fp':0.15, #Fraction of electrical power lost as heat to ambient
            'Wdot_max': 0.85*sqrt(3)*18*230, #W
            'Vdot_ratio': 1.0, #Displacement Scale factor
            }
    
        Comp=VICompressorLumpkinClass(**kwds)
        Comp.Calculate() 
        
        data_calc = {'Tdis':[Comp.Tout_r],
                    'mdot':[Comp.mdot_r],
                    'mdot_inj':[Comp.mdot_inj], 
                    'Wdot':[Comp.W],
                    'etaoa':[Comp.eta_oi],
                    'fq':[Comp.fq]} 
            
        
        # Write to Excel
        filename = os.path.dirname(__file__)+'\VIScroll_output.xlsx'
        xl = pd.read_excel(filename, sheet_name='Validation')

        df = pd.DataFrame(data=data_calc)

        df.reindex(columns=xl.columns)
        df_final=xl.append(df,ignore_index=True)
        df_final.tail()
        
        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl',index=False)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df_final.to_excel(writer,index=False,sheet_name='Validation')
        
        # 
        writer.save()    
    
    
def Validate_TwoPhaseInjection():
    
    # Read excel file scroll compressorr
    filename = os.path.dirname(__file__)+'\VIScroll_output.xlsx'
    df_meas = pd.read_excel(open(filename,'rb'), sheet_name='Experimental_Data_2phase',skip_footer=0)

    Tamb = np.asarray(df_meas['Tamb'].values.T[:].tolist()) #[K]
    Tsuc = np.asarray(df_meas['Tsuc'].values.T[:].tolist()) #[K]
    psuc = np.asarray(df_meas['Pev'].values.T[:].tolist())*1000 #[Pa]
    Tdis = np.asarray(df_meas['Tdis'].values.T[:].tolist()) #[K]
    pdis = np.asarray(df_meas['Pcond'].values.T[:].tolist())*1000 #[Pa]
    Tinj = np.asarray(df_meas['Tinj'].values.T[:].tolist()) #[K]
    pinj = np.asarray(df_meas['Pinj'].values.T[:].tolist())*1000 #[Pa]

    for i in range(len(Tamb)):
    
        #Abstract State        
        Ref = 'R407C'
        Backend = 'REFPROP' #choose between: 'HEOS','TTSE&HEOS','BICUBIC&HEOS','REFPROP','SRK','PR'
        AS = CP.AbstractState(Backend, Ref)
        

        kwds={
            'P':[2.671,1,0.06768,0.08695,0.4397,0.4899,-0.6598,0.1848,0.699,0.3221],
            'R':[0.08758,1,-1.97,0.5076,-2.745,1.334,-1.47,0.6262,0.3004,-0.2136],
            'T':[1.206,1,-0.01493,-0.005106,-0.5655,0.4426,-0.7388,-0.2694,0.2932,1.474],
            'V':[1.169,1,0.00497,-0.02236,-0.6345,0.3229,-0.1655,-0.3987,0.2806,1.856],
            'IS':[0.4,1,0.02304,-0.3125,3.792,0.5774,0.3489,-0.4794,-0.07093,1.974],
            'F':[10,1,0.3288,-1.248,6.196,0.9401,2.269,0.3922,0.4534,-2.865],
            'AS':AS,
            'Tin_r':Tsuc[i],
            'pin_r':psuc[i],
            'pout_r':pdis[i],
            'pinj_r':pinj[i],
            'Tinj_r':Tinj[i],
            'Tamb':Tamb[i],
            'Vdisp': 67.186962e-6, #[m3/rev]
            'f_power': 60, #[Hz]
            'f_nominal': 60, #[Hz]
            'fp':0.15, #Fraction of electrical power lost as heat to ambient
            'Wdot_max': 0.85*sqrt(3)*18*230, #W
            'Vdot_ratio': 1.0, #Displacement Scale factor
            }
    
        Comp=VICompressorLumpkinClass(**kwds)
        Comp.Calculate() 
        
        data_calc = {'Tdis':[Comp.Tout_r],
                    'mdot':[Comp.mdot_r],
                    'mdot_inj':[Comp.mdot_inj], 
                    'Wdot':[Comp.W],
                    'etaoa':[Comp.eta_oi],
                    'fq':[Comp.fq]} 
        
        print(Comp.Tout_r)    
        
        # Write to Excel
        filename = os.path.dirname(__file__)+'\VIScroll_output.xlsx'
        xl = pd.read_excel(filename, sheet_name='BiPI_Validation_2phase')

       #   df = pd.DataFrame(data=data_calc)

       #   df.reindex(columns=xl.columns)
        df_final=xl.append(df,ignore_index=True)
        df_final.tail()
        
        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl',index=False)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df_final.to_excel(writer,index=False,sheet_name='BiPI_Validation_2phase')
        
        # 
        writer.save()    

        

if __name__=='__main__':        



    #Validate_VaporInjection()
    Validate_TwoPhaseInjection()